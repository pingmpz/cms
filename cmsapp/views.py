from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.http import HttpResponse
import requests
from dateutil import parser
import json
# Authenticate
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
# File Reader
from openpyxl import load_workbook, Workbook
# Date Time
from datetime import datetime, timedelta
import time
from calendar import monthrange
# EMAIL
from django.core.mail import EmailMessage
from cms.settings import EMAIL_HOST_USER
import smtplib
import traceback
import threading
from django.template.loader import get_template
# File
from django.core.files.storage import FileSystemStorage
from pathlib import Path
import glob, os, shutil
# Token Generator
import secrets
# Line Noti
from django_line_notification.line_notify import Line

import random

from .models import SectionGroup, Employee, MachineGroup, Machine, Task, Vendor, Category, SubCategory, MailGroup, CriticalPart, SplindlePart, PasswordStorage, PasswordItem, Request, File, Member, RequestVendor, Comment, RequestSubCategory, OperatorWorkingTime, VendorWorkingTime, MachineDowntime, Costing, TotalOperationTime, QualityObjectiveTarget, EstimateWorkingTime

HOST_URL = 'http://129.1.100.185:8200/'
TEMPLATE_REQUEST = 'email_templates/request.html'
NEW_PV_REQUEST_MA_SIZE = 20

################################# Authenticate #################################

def first_page(request):
    sgs = SectionGroup.objects.all()
    context = {
        'sgs': sgs,
    }
    return render(request, 'first_page.html', context)

def track_request(request, search_text):
    reqs = []
    files = []
    if search_text != '0':
        # mcs = Machine.objects.filter(mc_no__icontains=search_text) | Machine.objects.filter(sap_mc_no__icontains=search_text) | Machine.objects.filter(section__icontains=search_text)
        # reqs = Request.objects.filter(type="User Request",emp_id=search_text).order_by('-request_date') | Request.objects.filter(type="User Request",mc__in=mcs).order_by('-request_date')
        reqs = Request.objects.filter(type="User Request")
        reqs = reqs.filter(emp_id=search_text) | reqs.filter(mc__mc_no__icontains=search_text) | reqs.filter(type="User Request",sg__name=search_text)
        reqs = reqs.order_by('-request_date')
        reqs = reqs[:100]
        for req in reqs:
            xfiles = File.objects.filter(req=req)
            if len(xfiles) >= 1:
                xfile = xfiles[0]
                files.append(xfile.file_name)
            else:
                files.append("")
    else:
        search_text = ""
    context = {
        'search_text': search_text,
        'reqs': reqs,
        'files': files,
    }
    return render(request, 'track_request.html', context)

def login_page(request):
    context = {
    }
    return render(request, 'login_page.html', context)

def validate_login(request):
    canLogIn = False
    username = request.GET['username']
    password = request.GET['password']
    user = authenticate(request, username=username, password=password)
    if user is not None:
        canLogIn = True
    data = {
        'canLogIn': canLogIn,
    }
    return JsonResponse(data)

def login_action(request):
    username = request.POST['username']
    password = request.POST['password']
    user = authenticate(request, username=username, password=password)
    if user is not None:
        login(request, user)
        return redirect('/index/')
    return redirect('/')

@login_required(login_url='/')
def logout_action(request):
    logout(request)
    return redirect('/')

@login_required(login_url='/')
def setting(request):
    # update_machine()
    users = []
    set_user = []
    if request.user.is_superuser or request.user.is_staff:
        users = sort_user_by_section(User.objects.all())
        set_user = get_set_user(users)
    # emps = Employee.objects.all()
    # for emp in emps:
    #     if emp.section == 'S-STDH':
    #         emp.section = 'STDH'
    #         emp.save()
    context = {
        'users': users,
        'set_user': set_user,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'setting.html', context)

##################################### Page #####################################

def new_request(request, sg_name):
    sgs = SectionGroup.objects.all()
    token = secrets.token_urlsafe(16)
    context = {
        'sg_name': sg_name,
        'sgs': sgs,
        'token': token,
    }
    return render(request, 'new_request.html', context)

def new_request_success(request, request_no):
    context = {
        'request_no': request_no,
    }
    return render(request, 'new_request_success.html', context)

def new_pv_request(request):
    sgs = SectionGroup.objects.all()
    mcs = Machine.objects.filter(is_active=True).order_by('section')
    tasks = Task.objects.filter(is_active=True).order_by('type')
    set_mc = get_set_mc(mcs)
    set_task = get_set_task(tasks)
    token = secrets.token_urlsafe(16)
    context = {
        'sgs': sgs,
        'mcs': mcs,
        'tasks': tasks,
        'set_mc': set_mc,
        'set_task': set_task,
        'token': token,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'new_pv_request.html', context)

def new_pv_request_ma(request):
    mcs = Machine.objects.filter(is_active=True).order_by('section')
    set_mc = get_set_mc(mcs)
    size = [''] * NEW_PV_REQUEST_MA_SIZE
    context = {
        'mcs': mcs,
        'set_mc': set_mc,
        'size': size,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'new_pv_request_ma.html', context)

@login_required(login_url='/')
def request_page(request, request_no):
    req_is_exist = Request.objects.filter(req_no=request_no).exists()
    req = None
    is_member = False
    members = []
    files = []
    req_vens = []
    comments = []
    req_sub_cats = []
    set_rsc = []
    files = []
    owts = []
    vwts = []
    mcdts = []
    costs = []
    total_cost = 0
    wt_len = 0

    sgs = []
    mcs = []
    users = []
    vens = []
    sub_cats = []

    set_mc = []
    set_user = []
    set_sc = []

    selected_members = []
    selected_vendors = []
    selected_sc = []
    if req_is_exist:
        req = Request.objects.get(req_no=request_no)
        is_member = Member.objects.filter(req=req,user=request.user).exists()
        members = Member.objects.filter(req=req)
        files = File.objects.filter(req=req)
        req_vens = RequestVendor.objects.filter(req=req)
        comments = Comment.objects.filter(req=req).order_by('-date_published')
        req_sub_cats = RequestSubCategory.objects.filter(req=req)
        set_rsc = get_set_rsc(req_sub_cats)
        files = File.objects.filter(req=req)
        owts = OperatorWorkingTime.objects.filter(req=req).order_by('-start_datetime')
        vwts = VendorWorkingTime.objects.filter(req=req).order_by('-start_datetime')
        mcdts = MachineDowntime.objects.filter(req=req).order_by('-start_datetime')
        costs = Costing.objects.filter(req=req).order_by('-date_published')
        total_cost = get_total_cost(costs)
        wt_len = len(owts) + len(vwts)
        sgs = SectionGroup.objects.all()
        if req.status == 'Pending': # Reduce Load
            mcs = Machine.objects.filter(is_active=True).order_by('section')
        users = sort_user_by_section(User.objects.filter(is_active=True))
        vens = Vendor.objects.filter(is_active=True)
        sub_cats = SubCategory.objects.all().order_by('cat')
        set_mc = get_set_mc(mcs)
        set_user = get_set_user(users)
        set_sc = get_set_sc(sub_cats)
        selected_members = get_selected_members(req, users)
        selected_vendors = get_selected_vendors(req, vens)
        selected_sc = get_selected_sc(req, sub_cats)
    context = {
        'request_no': request_no,
        'req_is_exist': req_is_exist,
        'req': req,
        'is_member': is_member,
        'members': members,
        'files': files,
        'req_vens': req_vens,
        'comments': comments,
        'req_sub_cats': req_sub_cats,
        'set_rsc': set_rsc,
        'files': files,
        'owts': owts,
        'vwts': vwts,
        'mcdts': mcdts,
        'costs': costs,
        'total_cost': total_cost,
        'wt_len': wt_len,
        'sgs': sgs,
        'mcs': mcs,
        'users': users,
        'vens': vens,
        'sub_cats': sub_cats,
        'set_mc': set_mc,
        'set_user': set_user,
        'set_sc': set_sc,
        'selected_members': selected_members,
        'selected_vendors': selected_vendors,
        'selected_sc': selected_sc,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'request_page.html', context)

def req(request, request_id):
    req_no = create_req_no(request_id)
    return redirect('/request_page/' + req_no)

#------------------------------------ Main ------------------------------------#

def all_page_data(request):
    is_in = is_in_section_group(request)
    my_reqs = []
    temp_reqs = Request.objects.filter(status='In Progress') | Request.objects.filter(status='On Hold')
    for req in temp_reqs:
        is_member = Member.objects.filter(req=req,user=request.user).exists()
        if is_member:
            my_reqs.append(req)
    my_request_count = len(my_reqs)

    pending_reqs = []
    if is_in_section_group(request):
        pending_reqs = Request.objects.filter(status='Pending',type='User Request',sg=request.user.employee.section)
    else :
        pending_reqs = Request.objects.filter(status='Pending',type='User Request')
    pending_request_count = len(pending_reqs)

    pv_pending_reqs = []
    if is_in_section_group(request):
        pv_pending_reqs = Request.objects.filter(status='Pending',type='Preventive',sg=request.user.employee.section)
    else :
        pv_pending_reqs = Request.objects.filter(status='Pending',type='Preventive')
    pv_pending_request_count = len(pv_pending_reqs)

    all_reqs = []
    if is_in_section_group(request):
        all_reqs = Request.objects.filter(status='In Progress',sg=request.user.employee.section) | Request.objects.filter(status='On Hold',sg=request.user.employee.section)
    else :
        all_reqs = Request.objects.filter(status='In Progress') | Request.objects.filter(status='On Hold')
    all_request_count = len(all_reqs)

    breakdown_reqs = []
    if is_in_section_group(request):
        breakdown_reqs = Request.objects.filter(is_breakdown=True,sg=request.user.employee.section)
    else :
        breakdown_reqs = Request.objects.filter(is_breakdown=True)
    breakdown_request_count = len(breakdown_reqs)

    cps = CriticalPart.objects.all()
    is_not_enough_cp = get_is_not_enough_cp(cps)
    sps = SplindlePart.objects.all()
    is_not_enough_sp = get_is_not_enough_sp(sps)
    context = {
        'is_in': is_in,
        'my_request_count': my_request_count,
        'pending_request_count': pending_request_count,
        'pv_pending_request_count': pv_pending_request_count,
        'all_request_count': all_request_count,
        'breakdown_request_count': breakdown_request_count,
        'is_not_enough_cp': is_not_enough_cp,
        'is_not_enough_sp': is_not_enough_sp,
    }
    return context

@login_required(login_url='/')
def index(request):
    reqs = []
    temp_reqs = Request.objects.filter(status='In Progress') | Request.objects.filter(status='On Hold')
    for req in temp_reqs:
        is_member = Member.objects.filter(req=req,user=request.user).exists()
        if is_member:
            reqs.append(req)
    has_wts = get_has_wts(reqs)
    has_mcdts = get_has_mcdts(reqs)
    context = {
        'reqs': reqs,
        'has_wts': has_wts,
        'has_mcdts': has_mcdts,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'index.html', context)

@login_required(login_url='/')
def request_pending(request, fsg):
    sgs = SectionGroup.objects.all()
    reqs = []
    if fsg == 'MY' and is_in_section_group(request):
        fsg = request.user.employee.section
    elif fsg == 'MY':
        fsg = 'ALL'

    reqs = Request.objects.filter(status='Pending',type='User Request')
    if fsg != 'ALL':
        reqs = reqs.filter(sg=fsg)

    context = {
        'sgs': sgs,
        'fsg': fsg,
        'reqs': reqs,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'request_pending.html', context)

@login_required(login_url='/')
def request_pv_pending(request, fsg):
    sgs = SectionGroup.objects.all()
    reqs = []
    if fsg == 'MY' and is_in_section_group(request):
        fsg = request.user.employee.section
    elif fsg == 'MY':
        fsg = 'ALL'

    reqs = Request.objects.filter(status='Pending',type='Preventive')
    if fsg != 'ALL':
        reqs = reqs.filter(sg=fsg)

    context = {
        'sgs': sgs,
        'fsg': fsg,
        'reqs': reqs,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'request_pv_pending.html', context)

@login_required(login_url='/')
def request_all(request, fsg, fstatus, ftype):
    sgs = SectionGroup.objects.all()
    reqs = []
    if fsg == 'MY' and is_in_section_group(request):
        fsg = request.user.employee.section
    elif fsg == 'MY':
        fsg = 'ALL'
    status = None
    if fstatus != 'ALL':
        status = fstatus.title()
    type = 'ALL'
    if ftype == 'PV':
        type = 'Preventive'
    elif ftype == 'UR':
        type = 'User Request'

    reqs = Request.objects.filter(status='In Progress')  | Request.objects.filter(status='On Hold')

    if fsg != 'ALL':
        reqs = reqs.filter(sg=fsg)
    if fstatus != 'ALL':
        reqs = reqs.filter(status=status)
    if ftype != 'ALL':
        reqs = reqs.filter(type=type)

    is_members = get_is_members(reqs, request)
    has_wts = get_has_wts(reqs)
    has_mcdts = get_has_mcdts(reqs)

    context = {
        'sgs': sgs,
        'fsg': fsg,
        'fstatus': fstatus,
        'ftype': ftype,
        'reqs': reqs,
        'is_members': is_members,
        'has_wts': has_wts,
        'has_mcdts': has_mcdts,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'request_all.html', context)

@login_required(login_url='/')
def request_breakdown(request, fsg):
    sgs = SectionGroup.objects.all()
    reqs = []
    if fsg == 'MY' and is_in_section_group(request):
        fsg = request.user.employee.section
    elif fsg == 'MY':
        fsg = 'ALL'

    reqs = Request.objects.filter(is_breakdown=True)
    if fsg != 'ALL':
        reqs = reqs.filter(sg=fsg)

    is_members = get_is_members(reqs, request)
    has_wts = get_has_wts(reqs)
    has_mcdts = get_has_mcdts(reqs)
    context = {
        'sgs': sgs,
        'fsg': fsg,
        'reqs': reqs,
        'is_members': is_members,
        'has_wts': has_wts,
        'has_mcdts': has_mcdts,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'request_breakdown.html', context)

@login_required(login_url='/')
def request_history(request, fsg, fstatus, ftype, fstartdate, fstopdate):
    sgs = SectionGroup.objects.all()
    if fsg == 'MY' and is_in_section_group(request):
        fsg = request.user.employee.section
    elif fsg == 'MY':
        fsg = 'ALL'
    status = fstatus.capitalize()
    type = 'ALL'
    if ftype == 'PV':
        type = 'Preventive'
    elif ftype == 'UR':
        type = 'User Request'
    if fstartdate == "LASTWEEK":
        fstartdate = (datetime.today() - timedelta(days=7)).strftime('%Y-%m-%d')
    if fstopdate == "TODAY":
        fstopdate = datetime.today().strftime('%Y-%m-%d')

    reqs = Request.objects.filter(status='Rejected',request_date__range=[fstartdate, fstopdate]) | Request.objects.filter(status='Complete',request_date__range=[fstartdate, fstopdate]) | Request.objects.filter(status='Canceled',request_date__range=[fstartdate, fstopdate])
    if fsg != 'ALL':
        reqs = reqs.filter(sg=fsg)
    if fstatus != 'ALL':
        reqs = reqs.filter(status=status)
    if ftype != 'ALL':
        reqs = reqs.filter(type=type)

    is_members = get_is_members(reqs, request)
    has_wts = get_has_wts(reqs)
    has_mcdts = get_has_mcdts(reqs)
    context = {
        'sgs': sgs,
        'fsg': fsg,
        'fstatus': fstatus,
        'ftype': ftype,
        'fstartdate': fstartdate,
        'fstopdate': fstopdate,
        'reqs': reqs,
        'is_members': is_members,
        'has_wts': has_wts,
        'has_mcdts': has_mcdts,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'request_history.html', context)

@login_required(login_url='/')
def cri_part_list(request):
    cps = CriticalPart.objects.all()
    context = {
        'cps': cps,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'cri_part_list.html', context)

@login_required(login_url='/')
def spl_part_list(request):
    sps = SplindlePart.objects.all()
    context = {
        'sps': sps,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'spl_part_list.html', context)

@login_required(login_url='/')
def pwst_list(request):
    pwsts = None
    pw_passwords, pw_dates = [], []
    if request.user.employee.enable_pwst:
        pwsts = PasswordStorage.objects.all()
        for pwst in pwsts:
            pwis = PasswordItem.objects.filter(pwst=pwst).order_by('-id')
            if pwis:
                pw_passwords.append(pwis[0].password)
                pw_dates.append(pwis[0].date_published)
            else:
                pw_passwords.append(None)
                pw_dates.append(None)
    context = {
        'pwsts': pwsts,
        'pw_passwords': pw_passwords,
        'pw_dates': pw_dates,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'pwst_list.html', context)

@login_required(login_url='/')
def mc_page(request, mc_no):
    mc = Machine.objects.get(mc_no=mc_no)
    history_reqs = Request.objects.filter(mc=mc,status='Complete') | Request.objects.filter(mc=mc,status='Canceled')
    active_reqs = Request.objects.filter(mc=mc,status='In Progress') | Request.objects.filter(mc=mc,status='On Hold')
    pending_reqs = Request.objects.filter(mc=mc,status='Pending')
    context = {
        'mc': mc,
        'history_reqs': history_reqs,
        'active_reqs': active_reqs,
        'pending_reqs': pending_reqs,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'mc_page.html', context)

#----------------------------------- Report -----------------------------------#

@login_required(login_url='/')
def summary(request, fsg):
    sgs = SectionGroup.objects.all()

    # REQUEST COUNTER
    all_req_count = 0
    pending_req_count = 0
    rejected_req_count = 0
    on_process_req_count = 0
    on_hold_req_count = 0
    complete_req_count = 0
    canceled_req_count = 0
    active_req_count = 0
    inactive_req_count = 0

    all_us_req_count = 0
    pending_us_req_count = 0
    rejected_us_req_count = 0
    on_process_us_req_count = 0
    on_hold_us_req_count = 0
    complete_us_req_count = 0
    canceled_us_req_count = 0
    active_us_req_count = 0
    inactive_us_req_count = 0

    all_pv_req_count = 0
    pending_pv_req_count = 0
    rejected_pv_req_count = 0
    on_process_pv_req_count = 0
    on_hold_pv_req_count = 0
    complete_pv_req_count = 0
    canceled_pv_req_count = 0
    active_pv_req_count = 0
    inactive_pv_req_count = 0
    if fsg == 'MY' and is_in_section_group(request):
        fsg = request.user.employee.section
    elif fsg == 'MY':
        fsg = 'ALL'
    if fsg == 'ALL':
        all_us_req_count = Request.objects.filter(type='User Request').count()
        all_pv_req_count = Request.objects.filter(type='Preventive').count()
        all_req_count = all_us_req_count + all_pv_req_count
        pending_us_req_count = Request.objects.filter(status='Pending',type='User Request').count()
        pending_pv_req_count = Request.objects.filter(status='Pending',type='Preventive').count()
        pending_req_count = pending_us_req_count + pending_pv_req_count
        rejected_us_req_count = Request.objects.filter(status='Rejected',type='User Request').count()
        rejected_pv_req_count = Request.objects.filter(status='Rejected',type='Preventive').count()
        rejected_req_count = rejected_us_req_count + rejected_pv_req_count
        on_process_us_req_count = Request.objects.filter(status='In Progress',type='User Request').count()
        on_process_pv_req_count = Request.objects.filter(status='In Progress',type='Preventive').count()
        on_process_req_count = on_process_us_req_count + on_process_pv_req_count
        on_hold_us_req_count = Request.objects.filter(status='On Hold',type='User Request').count()
        on_hold_pv_req_count = Request.objects.filter(status='On Hold',type='Preventive').count()
        on_hold_req_count = on_hold_us_req_count + on_hold_pv_req_count
        complete_us_req_count = Request.objects.filter(status='Complete',type='User Request').count()
        complete_pv_req_count = Request.objects.filter(status='Complete',type='Preventive').count()
        complete_req_count = complete_us_req_count + complete_pv_req_count
        canceled_us_req_count = Request.objects.filter(status='Canceled',type='User Request').count()
        canceled_pv_req_count = Request.objects.filter(status='Canceled',type='Preventive').count()
        canceled_req_count = canceled_us_req_count + canceled_pv_req_count
    else:
        all_us_req_count = Request.objects.filter(type='User Request',sg=fsg).count()
        all_pv_req_count = Request.objects.filter(type='Preventive',sg=fsg).count()
        all_req_count = all_us_req_count + all_pv_req_count
        pending_us_req_count = Request.objects.filter(status='Pending',type='User Request',sg=fsg).count()
        pending_pv_req_count = Request.objects.filter(status='Pending',type='Preventive',sg=fsg).count()
        pending_req_count = pending_us_req_count + pending_pv_req_count
        rejected_us_req_count = Request.objects.filter(status='Rejected',type='User Request',sg=fsg).count()
        rejected_pv_req_count = Request.objects.filter(status='Rejected',type='Preventive',sg=fsg).count()
        rejected_req_count = rejected_us_req_count + rejected_pv_req_count
        on_process_us_req_count = Request.objects.filter(status='In Progress',type='User Request',sg=fsg).count()
        on_process_pv_req_count = Request.objects.filter(status='In Progress',type='Preventive',sg=fsg).count()
        on_process_req_count = on_process_us_req_count + on_process_pv_req_count
        on_hold_us_req_count = Request.objects.filter(status='On Hold',type='User Request',sg=fsg).count()
        on_hold_pv_req_count = Request.objects.filter(status='On Hold',type='Preventive',sg=fsg).count()
        on_hold_req_count = on_hold_us_req_count + on_hold_pv_req_count
        complete_us_req_count = Request.objects.filter(status='Complete',type='User Request',sg=fsg).count()
        complete_pv_req_count = Request.objects.filter(status='Complete',type='Preventive',sg=fsg).count()
        complete_req_count = complete_us_req_count + complete_pv_req_count
        canceled_us_req_count = Request.objects.filter(status='Canceled',type='User Request',sg=fsg).count()
        canceled_pv_req_count = Request.objects.filter(status='Canceled',type='Preventive',sg=fsg).count()
        canceled_req_count = canceled_us_req_count + canceled_pv_req_count
    active_us_req_count = pending_us_req_count + on_process_us_req_count + on_hold_us_req_count
    active_pv_req_count = pending_pv_req_count + on_process_pv_req_count + on_hold_pv_req_count
    active_req_count = active_us_req_count + active_pv_req_count
    inactive_us_req_count = rejected_us_req_count + complete_us_req_count + canceled_us_req_count
    inactive_pv_req_count = rejected_pv_req_count + complete_pv_req_count + canceled_pv_req_count
    inactive_req_count = inactive_us_req_count + inactive_pv_req_count

    days = 7
    i = 0
    cat_data = [""] * days
    new_data = [0] * days
    complete_ur_data = [0] * days
    complete_pv_data = [0] * days
    while i < days:
        date = (datetime.today() - timedelta(days=(days - i))).strftime('%Y-%m-%d')
        cat_data[i] = (datetime.today() - timedelta(days=(days - i))).strftime('%d %b')
        if fsg == 'ALL':
            new_data[i] = int(Request.objects.filter(type='User Request',date_published__date=date).count())
            complete_ur_data[i] = int(Request.objects.filter(type='User Request',finish_datetime__date=date).count())
            complete_pv_data[i] = int(Request.objects.filter(type='Preventive',finish_datetime__date=date).count())
        else:
            new_data[i] = int(Request.objects.filter(status='Complete',type='User Request',date_published__date=date,sg=fsg).count())
            complete_ur_data[i] = int(Request.objects.filter(status='Complete',type='User Request',finish_datetime__date=date,sg=fsg).count())
            complete_pv_data[i] = int(Request.objects.filter(status='Complete',type='Preventive',finish_datetime__date=date,sg=fsg).count())
        i = i + 1

    context = {
        'sgs': sgs,
        'fsg': fsg,
        'all_req_count': all_req_count,
        'pending_req_count': pending_req_count,
        'rejected_req_count': rejected_req_count,
        'on_process_req_count': on_process_req_count,
        'on_hold_req_count': on_hold_req_count,
        'complete_req_count': complete_req_count,
        'canceled_req_count': canceled_req_count,
        'active_req_count': active_req_count,
        'inactive_req_count': inactive_req_count,

        'all_us_req_count': all_us_req_count,
        'pending_us_req_count': pending_us_req_count,
        'rejected_us_req_count': rejected_us_req_count,
        'on_process_us_req_count': on_process_us_req_count,
        'on_hold_us_req_count': on_hold_us_req_count,
        'complete_us_req_count': complete_us_req_count,
        'canceled_us_req_count': canceled_us_req_count,
        'active_us_req_count': active_us_req_count,
        'inactive_us_req_count': inactive_us_req_count,

        'all_pv_req_count': all_pv_req_count,
        'pending_pv_req_count': pending_pv_req_count,
        'rejected_pv_req_count': rejected_pv_req_count,
        'on_process_pv_req_count': on_process_pv_req_count,
        'on_hold_pv_req_count': on_hold_pv_req_count,
        'complete_pv_req_count': complete_pv_req_count,
        'canceled_pv_req_count': canceled_pv_req_count,
        'active_pv_req_count': active_pv_req_count,
        'inactive_pv_req_count': inactive_pv_req_count,

        'cat_data': cat_data,
        'new_data': new_data,
        'complete_ur_data': complete_ur_data,
        'complete_pv_data': complete_pv_data,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'report/summary.html', context)

@login_required(login_url='/')
def pv_calendar(request, fmcg, fyear):
    mcgs = MachineGroup.objects.all()
    years = get_years()
    sg = SectionGroup.objects.get(name='MA')
    reqs = []
    if fyear == 'THISYEAR':
        fyear = datetime.today().strftime('%Y')
    reqs = Request.objects.filter(type='Preventive',sg=sg,request_date__year=fyear).exclude(status='Canceled').exclude(status='Rejected').order_by('mc') 
    mcs = []
    mcs_months = []
    months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
    nmonths = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12']
    for req in reqs:
        if req.mc and req.mc not in mcs:
            if fmcg == 'ALL':
                mcs.append(req.mc)
            if fmcg == 'NONE':
                if req.mc.mcg == None:
                    mcs.append(req.mc)
            else:
                mcg = MachineGroup.objects.get(id=fmcg)
                if req.mc.mcg == mcg:
                    mcs.append(req.mc)
    mc_nos = []
    for mc in mcs:
        mc_nos.append(mc.mc_no)
        xreqs = reqs.filter(mc=mc)
        mc_months = []
        for nmonth in nmonths:
            freqs = xreqs.filter(request_date__month=nmonth)
            if freqs:
                mc_months.append('Y')
            else:
                mc_months.append('N')
        mcs_months.append(mc_months)
    context = {
        'mcgs': mcgs,
        'years': years,
        'fmcg': fmcg,
        'fyear': fyear,
        'mc_nos': mc_nos,
        'mcs_months': mcs_months,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'report/pv_calendar.html', context)

@login_required(login_url='/')
def report_q_obj(request, fmcg, fyear):
    mcgs = MachineGroup.objects.all()
    if fmcg == 'FIRST':
        fmcg = mcgs[0].id
    mcg = MachineGroup.objects.get(id=fmcg)
    years = get_years()
    if fyear == 'THISYEAR':
        fyear = datetime.today().strftime('%Y')

    sg = SectionGroup.objects.get(name='MA')
    mcs = Machine.objects.filter(mcg=mcg)
    months = ['January', 'February', 'March', 'April', 'May', 'June', 'July','August', 'September', 'October', 'November', 'December']
    req_count = []
    tot_mins = []
    tot_hrs = []
    ewt_hrs = []
    dt_mins = []
    dt_hrs = []
    dt_pers = []
    target_dts = []
    mttrs = []
    target_mttrs = []
    mtbfs = []
    target_mtbfs = []
    mcdt_cals = [] * len(months)
    min_cals = [] * len(months)
    hr_cals = [] * len(months)
    for month in months:
        month_no = months.index(month) + 1
        # No of Request
        reqs = []
        temp_reqs = Request.objects.filter(mc__in=mcs,type='User Request',status='Complete',sg=sg,request_date__year=fyear,request_date__month=month_no)
        # Only Request with Downtime
        for req in temp_reqs:
            mcdt_is_exist = MachineDowntime.objects.filter(req=req).exists()
            if mcdt_is_exist:
                reqs.append(req)
        req_count.append(len(reqs))
        # TotalOperationTime
        tot_min = 0
        tot_hr = 0
        tot_is_exist = TotalOperationTime.objects.filter(mcg=mcg,year=fyear,month=month_no).exists()
        if tot_is_exist:
            tot = TotalOperationTime.objects.get(mcg=mcg,year=fyear,month=month_no)
            tot_min = int(tot.time)
            tot_hr = int(tot.time / 60)
        tot_mins.append(tot_min)
        tot_hrs.append(tot_hr)
        # EstimateWorkingTime
        ewt_hr = get_ewt(mcg, fyear, month_no)
        ewt_hrs.append(ewt_hr)
        # Downtime
        dt_min = 0
        dt_hr = 0
        dt_per = 0
        temp_mcdt_cals = []
        temp_min_cals = []
        temp_hr_cals = []
        for req in reqs:
            mcdts = MachineDowntime.objects.filter(req=req)
            for mcdt in mcdts:
                minutes_diff = (mcdt.stop_datetime - mcdt.start_datetime).total_seconds() / 60
                hours_diff = (mcdt.stop_datetime - mcdt.start_datetime).total_seconds() / 3600
                dt_min = dt_min + minutes_diff
                dt_hr = dt_hr + hours_diff
                temp_mcdt_cals.append(mcdt)
                temp_min_cals.append(minutes_diff)
                temp_hr_cals.append(hours_diff)
        mcdt_cals.append(temp_mcdt_cals)
        min_cals.append(temp_min_cals)
        hr_cals.append(temp_hr_cals)
        dt_min = int(dt_min)
        dt_hr = int(dt_hr)
        if ewt_hr != 0:
            dt_per = round(((dt_hr * 100) / ewt_hr), 2)
        dt_mins.append(dt_min)
        dt_hrs.append(dt_hr)
        dt_pers.append((dt_per))
        # Downtime Target
        target_dts.append(get_qot(mcg, 'DT', fyear, month_no))
        # MTTR
        if len(reqs):
            mttrs.append(int(dt_hr/len(reqs)))
        else:
            mttrs.append(0)
        target_mttrs.append(get_qot(mcg, 'MTTR', fyear, month_no))
        # MTBF
        if len(reqs):
            mtbfs.append(int(tot_hr/len(reqs)))
        else:
            mtbfs.append(0)
        target_mtbfs.append(get_qot(mcg, 'MTBF', fyear, month_no))
    context = {
        'mcgs': mcgs,
        'fmcg': fmcg,
        'mcg': mcg,
        'years': years,
        'fyear': fyear,
        'months': months,
        'req_count': req_count,
        'tot_mins': tot_mins,
        'tot_hrs': tot_hrs,
        'ewt_hrs': ewt_hrs,
        'dt_mins': dt_mins,
        'dt_hrs': dt_hrs,
        'dt_pers': dt_pers,
        'target_dts': target_dts,
        'mttrs': mttrs,
        'target_mttrs': target_mttrs,
        'mtbfs': mtbfs,
        'target_mtbfs': target_mtbfs,
        'mcdt_cals': mcdt_cals,
        'min_cals': min_cals,
        'hr_cals': hr_cals,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'report/q_obj.html', context)

@login_required(login_url='/')
def working_time(request, fuser, fstartdate, fstopdate):
    users = sort_user_by_section(User.objects.filter(is_active=True))
    set_user = get_set_user(users)
    user = None
    if fuser == "MY":
        user = request.user
        fuser = user.username
    else:
        user = User.objects.get(username=fuser)
    if fstartdate == "LASTWEEK":
        fstartdate = (datetime.today() - timedelta(days=7)).strftime('%Y-%m-%d')
    if fstopdate == "TODAY":
        fstopdate = datetime.today().strftime('%Y-%m-%d')
    d0 = datetime.strptime(fstartdate, '%Y-%m-%d')
    d1 = datetime.strptime(fstopdate, '%Y-%m-%d')
    days = (d1 - d0).days
    wt_data = [0] * (days + 1)
    cat_data = [0] * (days + 1)
    i = 0
    while i <= days:
        date = (datetime.strptime(fstopdate, '%Y-%m-%d') - timedelta(days=(days - i))).strftime('%Y-%m-%d')
        cat_data[i] = (datetime.strptime(fstopdate, '%Y-%m-%d') - timedelta(days=(days - i))).strftime('%d %b')
        time = 0
        wts = OperatorWorkingTime.objects.filter(user=user,start_datetime__date=date)
        for wt in wts:
            time = time + (wt.stop_datetime - wt.start_datetime).total_seconds() / 60
        wt_data[i] = time
        i = i + 1
    wts = OperatorWorkingTime.objects.filter(user=user,start_datetime__date__range=[fstartdate, fstopdate])
    durations = [] # Minute
    for wt in wts:
        durations.append(int((wt.stop_datetime - wt.start_datetime).total_seconds() / 60))
    context = {
        'fuser': fuser,
        'user': user,
        'users': users,
        'set_user': set_user,
        'fstartdate': fstartdate,
        'fstopdate': fstopdate,
        'cat_data': cat_data,
        'wt_data': wt_data,
        'wts': wts,
        'durations': durations,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'report/working_time.html', context)

@login_required(login_url='/')
def report_mc_dt(request, fsection, fmonth):
    if fmonth == 'THISMONTH':
        fmonth = datetime.today().strftime('%Y-%m')
    mcs = Machine.objects.filter(is_active=True).order_by('section')
    sections = get_set_mc(mcs)

    mcs = Machine.objects.filter(section=fsection)
    days = monthrange(int(fmonth[:4]), int(fmonth[5:7]))[1]
    dt_data = [0] * (days)
    cat_data = [0] * (days)
    i = 0
    #-- Only Start Date
    while i < days:
        date = fmonth + '-' + add_front_zero(str(i + 1))
        cat_data[i] = i + 1
        time = 0
        mcdts = MachineDowntime.objects.filter(start_datetime__date=date,mc__in=mcs)
        for mcdt in mcdts:
            time = time + (mcdt.stop_datetime - mcdt.start_datetime).total_seconds() / 60
        dt_data[i] = time
        i = i + 1
    mcdts = MachineDowntime.objects.filter(start_datetime__month=fmonth[5:7],start_datetime__year=fmonth[:4],mc__in=mcs)
    durations = [] # Minute
    for mcdt in mcdts:
        durations.append(int((mcdt.stop_datetime - mcdt.start_datetime).total_seconds() / 60))

    #-- 2 Average
    # while i < days:
    #     cat_data[i] = i + 1
    #     i = i + 1
    # mcdts = MachineDowntime.objects.filter(start_datetime__month=fmonth[5:7],start_datetime__year=fmonth[:4],mc__in=mcs)
    # for mcdt in mcdts:
    #     j = int(mcdt.start_datetime.day - 1)
    #     if mcdt.start_datetime.date() == mcdt.stop_datetime.date():
    #         time = (mcdt.stop_datetime - mcdt.start_datetime).total_seconds() / 60
    #         dt_data[j] = dt_data[j] + time
    #     else:
    #         day_diff = (mcdt.stop_datetime.date() - mcdt.start_datetime.date()).days
    #         time = ((mcdt.stop_datetime - mcdt.start_datetime).total_seconds() / 60)
    #         time = int(time / (day_diff + 1))
    #         k = j + day_diff
    #         while j <= k and j < days:
    #             dt_data[j] = dt_data[j] + time
    #             j = j + 1
    #-- 3 Full
    # while i < days:
    #     date = fmonth + '-' + add_front_zero(str(i + 1))
    #     cat_data[i] = i + 1
    #     time = 0
    #     mcdts = MachineDowntime.objects.filter(start_datetime__date=date,stop_datetime__date=date,mc__in=mcs)
    #     for mcdt in mcdts:
    #         time = time + (mcdt.stop_datetime - mcdt.start_datetime).total_seconds() / 60
    #     mcdts = MachineDowntime.objects.filter(start_datetime__date=date,mc__in=mcs).exclude(stop_datetime__date=date)
    #     for mcdt in mcdts:
    #         time = time + ( - mcdt.start_datetime).total_seconds() / 60
    #     dt_data[i] = time
    #     i = i + 1
    # mcdts = MachineDowntime.objects.filter(start_datetime__month=fmonth[5:7],start_datetime__year=fmonth[:4],mc__in=mcs)
    context = {
        'fsection' : fsection,
        'fmonth' : fmonth,
        'sections': sections,
        'cat_data': cat_data,
        'dt_data': dt_data,
        'mcdts': mcdts,
        'durations': durations,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'report/mc_dt.html', context)

@login_required(login_url='/')
def report_cat(request, fsc):
    sub_cats = SubCategory.objects.all().order_by('cat')
    set_sc = get_set_sc(sub_cats)
    sub_cat = None
    if fsc == 'FIRST':
        sub_cat = sub_cats[0]
    else:
        sub_cat = SubCategory.objects.get(id=fsc)
    rscs = RequestSubCategory.objects.filter(sub_cat=sub_cat)
    context = {
        'sub_cats': sub_cats,
        'set_sc': set_sc,
        'fsc': fsc,
        'sub_cat': sub_cat,
        'rscs': rscs,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'report/cat.html', context)

#----------------------------------- Master -----------------------------------#

@login_required(login_url='/')
def master_emp(request):
    users = User.objects.all()
    context = {
        'users': users,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'master/emp.html', context)

@login_required(login_url='/')
def master_mcg(request):
    mcgs = MachineGroup.objects.all()
    context = {
        'mcgs': mcgs,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'master/mcg.html', context)

@login_required(login_url='/')
def master_mc(request):
    mcs = Machine.objects.all()
    # for mc in mcs:
    #     if mc.plant == 'A' or mc.plant == 'E':
    #         mc.plant = 'CCSA'
    #         mc.save()
    #     elif mc.plant == 'P':
    #         mc.plant = 'CCSP'
    #         mc.save()
    context = {
        'mcs': mcs,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'master/mc.html', context)

@login_required(login_url='/')
def master_task(request):
    tasks = Task.objects.all()
    context = {
        'tasks': tasks,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'master/task.html', context)

@login_required(login_url='/')
def master_ven(request):
    vens = Vendor.objects.all()
    context = {
        'vens': vens,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'master/ven.html', context)

@login_required(login_url='/')
def master_cat(request):
    cats = Category.objects.all()
    context = {
        'cats': cats,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'master/cat.html', context)

@login_required(login_url='/')
def master_sub_cat(request):
    sub_cats = SubCategory.objects.all()
    context = {
        'sub_cats': sub_cats,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'master/sub_cat.html', context)

@login_required(login_url='/')
def master_sg(request):
    sgs = SectionGroup.objects.all()
    context = {
        'sgs': sgs,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'master/sg.html', context)

@login_required(login_url='/')
def master_mg(request):
    mgs = MailGroup.objects.all()
    custom_script()
    context = {
        'mgs': mgs,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'master/mg.html', context)

#---------------------------------- New Data ----------------------------------#

@login_required(login_url='/')
def new_emp(request):
    context = {
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'new/emp.html', context)

@login_required(login_url='/')
def new_mc(request):
    mcgs = MachineGroup.objects.all()
    context = {
        'mcgs': mcgs,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'new/mc.html', context)

@login_required(login_url='/')
def new_cp(request):
    context = {
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'new/cp.html', context)

@login_required(login_url='/')
def new_sp(request):
    context = {
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'new/sp.html', context)

@login_required(login_url='/')
def new_task(request):
    context = {
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'new/task.html', context)

@login_required(login_url='/')
def new_ven(request):
    context = {
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'new/ven.html', context)

@login_required(login_url='/')
def new_cat(request):
    context = {
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'new/cat.html', context)

@login_required(login_url='/')
def new_sub_cat(request):
    cats = Category.objects.all()
    context = {
        'cats': cats,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'new/sub_cat.html', context)

@login_required(login_url='/')
def new_mg(request):
    sgs = SectionGroup.objects.all()
    users = sort_user_by_section(User.objects.all())
    set_user = get_set_user(users)
    context = {
        'sgs': sgs,
        'users': users,
        'set_user': set_user,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'new/mg.html', context)

@login_required(login_url='/')
def new_pwst(request):
    context = {
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'new/pwst.html', context)

#--------------------------------- Edit Data ----------------------------------#

def edit_request(request,  request_no):
    req = Request.objects.get(req_no=request_no)
    mcs = Machine.objects.filter(is_active=True).order_by('section')
    set_mc = get_set_mc(mcs)
    token = secrets.token_urlsafe(16)
    context = {
        'request_no': request_no,
        'req': req,
        'mcs': mcs,
        'set_mc': set_mc,
        'token': token,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'edit_request.html', context)

@login_required(login_url='/')
def edit_mc(request, fmc):
    mcs = Machine.objects.all().order_by('section')
    set_mc = get_set_mc(mcs)
    if fmc == 'FIRST':
        fmc = mcs[0].mc_no
    mc = Machine.objects.get(mc_no=fmc)
    mcgs = MachineGroup.objects.all()
    context = {
        'fmc': fmc,
        'mcs': mcs,
        'set_mc': set_mc,
        'mc': mc,
        'mcgs': mcgs,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'edit/mc.html', context)

@login_required(login_url='/')
def edit_cp(request, fcp):
    cps = CriticalPart.objects.all().order_by('name')
    if fcp == 'FIRST':
        fcp = cps[0].id
    cp = CriticalPart.objects.get(id=fcp)
    context = {
        'fcp': fcp,
        'cps': cps,
        'cp': cp,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'edit/cp.html', context)

@login_required(login_url='/')
def edit_sp(request, fsp):
    sps = SplindlePart.objects.all()
    if fsp == 'FIRST':
        fsp = sps[0].id
    sp = SplindlePart.objects.get(id=fsp)
    context = {
        'fsp': fsp,
        'sps': sps,
        'sp': sp,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'edit/sp.html', context)

@login_required(login_url='/')
def edit_task(request, ftask):
    tasks = Task.objects.all().order_by('type')
    set_task = get_set_task(tasks)
    if ftask == 'FIRST':
        ftask = tasks[0].id
    task = Task.objects.get(id=ftask)
    context = {
        'ftask': ftask,
        'tasks': tasks,
        'set_task': set_task,
        'task': task,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'edit/task.html', context)

@login_required(login_url='/')
def edit_ven(request, fven):
    vens = Vendor.objects.all().order_by('code')
    if fven == 'FIRST':
        fven = vens[0].code
    ven = Vendor.objects.get(code=fven)
    context = {
        'fven': fven,
        'vens': vens,
        'ven': ven,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'edit/ven.html', context)

@login_required(login_url='/')
def edit_pwst(request, fpwst):
    pwsts = []
    pwst = None
    pwis = []
    if request.user.employee.enable_pwst:
        pwsts = PasswordStorage.objects.all().order_by('name')
        if fpwst == 'FIRST':
            fpwst = pwsts[0].id
        pwst = PasswordStorage.objects.get(id=fpwst)
        pwis = PasswordItem.objects.filter(pwst=pwst).order_by('-id')
    context = {
        'fpwst': fpwst,
        'pwsts': pwsts,
        'pwst': pwst,
        'pwis': pwis,
    }
    context['all_page_data'] = (all_page_data(request))
    return render(request, 'edit/pwst.html', context)

#################################### POST ######################################
def setting_save(request):
    is_reset = True if request.POST.get('is_reset', False) == 'on' else False
    new_password = request.POST['new_password']
    name = request.POST['name']
    section = request.POST['section']
    email = request.POST['email']
    phone_no = request.POST['phone_no']
    scheme = request.POST['scheme']
    sidebar = request.POST['sidebar']
    pv_created = request.POST['pv_created']
    auto_add = request.POST['auto_add']
    default_owt = request.POST['default_owt']
    reset_password_username = request.POST['reset_password_username'] if request.POST['reset_password_username'] != 'Select' else None
    user = request.user
    user.email = email
    if(is_reset):
        user.set_password(new_password)
    user.save()
    emp = request.user.employee
    emp.name = name
    emp.section = section
    emp.phone_no = phone_no
    emp.scheme = scheme
    emp.sidebar = sidebar
    emp.pv_created = pv_created
    emp.auto_add = auto_add
    emp.default_owt = default_owt
    emp.save()
    #-- Admin Function
    if reset_password_username != None:
        u = User.objects.get(username=reset_password_username)
        u.set_password('Ccs.1234')
        u.save()
    return redirect('/setting/')

def new_request_save(request):
    token = request.POST['token']
    emp_id = request.POST['emp_id'].strip()
    name = request.POST['name'].strip()
    section = request.POST['section'].strip()
    email = request.POST['email'].strip()
    phone_no = request.POST['phone_no'].strip()
    sg_name = request.POST['sg_name']
    description = request.POST['description']
    cc_to_me = True if request.POST.get('cc_to_me', False) == 'on' else False
    type = 'User Request'
    status = 'Pending'
    request_date = datetime.now().date()
    sg = SectionGroup.objects.get(name=sg_name)
    request_new = Request(emp_id=emp_id,name=name,section=section,email=email,phone_no=phone_no,sg=sg,type=type,status=status,request_date=request_date,description=description)
    request_new.save()
    request_new.req_no = create_req_no(request_new.id)
    request_new.save()
    #-- File Manage
    source_dir = 'media/temp/' + token
    target_dir = 'media/request/' + request_new.req_no
    try:
        os.listdir(source_dir)
        os.mkdir(target_dir)
    except:
        print('No File Upload')
    try:
        for file_name in os.listdir(source_dir):
            shutil.move(os.path.join(source_dir, file_name), os.path.join(target_dir))
            file_new = File(req=request_new,file_name=file_name)
            file_new.save()
        shutil.rmtree(source_dir, ignore_errors=False, onerror=None)
    except:
        print('Path Not Found')
    #-- Email
    subject = '[CMS] New Request #' + request_new.req_no
    send_to = get_mail_group(sg, False)
    cc_to = get_mail_group(sg, True)
    if cc_to_me:
        cc_to.append(request_new.email)
    email_template = get_template(TEMPLATE_REQUEST)
    email_content = email_template.render({
        'type' : 'New Request',
        'req' : request_new,
        'host_url' : HOST_URL,
    })
    send_email(subject, email_content, send_to, cc_to)
    #-- Line
    send_line(sg.line_token, request_new)
    return redirect('/new_request_success/' + request_new.req_no)

def new_pv_request_save(request):
    token = request.POST['token']
    sg_name = request.POST['sg_name']
    request_date = request.POST['req_date']
    description = request.POST['description']
    obj_type = request.POST['obj_type']
    mc_no = request.POST['mc_no'] if request.POST['obj_type'] == 'MACHINE' else None
    task_id = request.POST['task_id'] if request.POST['obj_type'] == 'TASK' else None
    mc = None
    if mc_no != None:
        mc = Machine.objects.get(mc_no=mc_no)
    task = None
    if task_id != None:
        task = Task.objects.get(id=task_id)
    type = 'Preventive'
    status = 'Pending'
    sg = SectionGroup.objects.get(name=sg_name)
    emp_id = request.user.username
    name = request.user.employee.name
    section = request.user.employee.section
    email = request.user.email
    phone_no = request.user.employee.phone_no
    request_new = Request(emp_id=emp_id,name=name,section=section,email=email,phone_no=phone_no,sg=sg,type=type,status=status,request_date=request_date,description=description,mc=mc,task=task)
    request_new.save()
    request_new.req_no = create_req_no(request_new.id)
    request_new.save()
    #-- File Manage
    source_dir = 'media/temp/' + token
    target_dir = 'media/request/' + request_new.req_no
    try:
        os.listdir(source_dir)
        os.mkdir(target_dir)
    except:
        print('No File Upload')
    try:
        for file_name in os.listdir(source_dir):
            shutil.move(os.path.join(source_dir, file_name), os.path.join(target_dir))
            file_new = File(req=request_new,file_name=file_name)
            file_new.save()
        shutil.rmtree(source_dir, ignore_errors=False, onerror=None)
    except:
        print('Path Not Found')
    #-- Navigate
    if request.user.employee.pv_created == 'Request Page':
        return redirect('/request_page/' + request_new.req_no)
    return redirect('/new_pv_request/')

def new_pv_request_ma_save(request):
    request_date = request.POST['req_date']
    mc_no_list, desc_list = [], []
    size = [''] * NEW_PV_REQUEST_MA_SIZE
    for idx, i in enumerate(size):
        mc_no = request.POST['mc_no_' + str(idx)] if request.POST['mc_no_' + str(idx)] != 'None' else None
        desc = (request.POST['desc_' + str(idx)]).strip()
        if mc_no != None:
            mc_no_list.append(mc_no)
            desc_list.append(desc)
    for idx, mc_no in enumerate(mc_no_list):
        task = None
        mc = Machine.objects.get(mc_no=mc_no)
        description = desc_list[idx]
        type = 'Preventive'
        status = 'Pending'
        sg = SectionGroup.objects.get(name='MA')
        emp_id = request.user.username
        name = request.user.employee.name
        section = request.user.employee.section
        email = request.user.email
        phone_no = request.user.employee.phone_no
        request_new = Request(emp_id=emp_id,name=name,section=section,email=email,phone_no=phone_no,sg=sg,type=type,status=status,request_date=request_date,description=description,mc=mc,task=task)
        request_new.save()
        request_new.req_no = create_req_no(request_new.id)
        request_new.save()
    return redirect('/new_pv_request_ma/')

def edit_request_save(request):
    token = request.POST['token']
    req_id = request.POST['req_id']
    request_date = request.POST['req_date']
    description = request.POST['description']
    mc_no = request.POST['mc_no'] if request.POST['mc_no'] != 'Select' else None
    mc = None
    is_same_mc = False
    req = Request.objects.get(id=req_id)
    if mc_no != None:
        mc = Machine.objects.get(mc_no=mc_no)
        if req.mc == mc:
            is_same_mc = True
    if not is_same_mc:
        req.is_breakdown = False
        mcdts = MachineDowntime.objects.filter(req=req)
        mcdts.delete()
    req.request_date = request_date
    req.mc = mc
    req.description = description
    req.save()
    #-- File Manage
    source_dir = 'media/temp/' + token
    target_dir = 'media/request/' + req.req_no
    try:
        os.listdir(source_dir)
        try:
            os.mkdir(target_dir)
        except:
            print('Folder is already Exist')
        try:
            for file_name in os.listdir(source_dir):
                shutil.move(os.path.join(source_dir, file_name), os.path.join(target_dir))
                file_new = File(req=req,file_name=file_name)
                file_new.save()
            shutil.rmtree(source_dir, ignore_errors=False, onerror=None)
        except:
            print('Path Not Found')
    except:
        print('No File Upload')
    return redirect('/request_page/' + req.req_no)

def new_emp_save(request):
    username = request.POST['new_username'].strip()
    password = request.POST['new_password']
    name = request.POST['name'].strip()
    section = request.POST['section'].strip()
    email = request.POST['email'].strip()
    phone_no = request.POST['phone_no'].strip()
    user_new = User.objects.create_user(username, '', password)
    user_new.email = email
    user_new.save()
    employee_new = Employee(user=user_new,name=name,section=section,phone_no=phone_no)
    employee_new.save()
    return redirect('/new/emp/')

def new_mc_save(request):
    mc_no = request.POST['mc_no'].strip()
    sap_mc_no = request.POST['sap_mc_no'].strip()
    type = request.POST['type'].strip()
    section = request.POST['section'].strip()
    mcg_id = request.POST['mcg_id']
    register_no = request.POST['register_no'].strip()
    asset_no = request.POST['asset_no'].strip()
    serial_no = request.POST['serial_no'].strip()
    manufacture = request.POST['manufacture'].strip()
    model = request.POST['model'].strip()
    plant = request.POST['plant'].strip()
    power = request.POST['power'].strip()
    mcg = None
    if mcg_id != '-1':
        mcg = MachineGroup.objects.get(id=mcg_id)
    location = request.POST['location'].strip()
    install_date = request.POST['install_date'] if request.POST['install_date'] != "" else None
    capacity = request.POST['capacity']
    note = request.POST['note']
    mc_new = Machine(mc_no=mc_no,sap_mc_no=sap_mc_no,type=type,section=section,mcg=mcg,register_no=register_no,asset_no=asset_no,serial_no=serial_no,manufacture=manufacture,model=model,plant=plant,power=power,location=location,install_date=install_date,capacity=capacity,note=note)
    mc_new.save()
    return redirect('/mc_page/' + str(mc_no))

def new_cp_save(request):
    name = request.POST['name'].strip()
    mat_code = request.POST['mat_code'].strip()
    amount = int(request.POST['amount'].strip())
    minimum = int(request.POST['minimum'].strip())
    note = request.POST['note']
    cp_new = CriticalPart(name=name,mat_code=mat_code,amount=amount,minimum=minimum,note=note)
    cp_new.save()
    return redirect('/new/cp/')

def new_sp_save(request):
    machine = request.POST['machine'].strip()
    model = request.POST['model'].strip()
    amount = int(request.POST['amount'].strip())
    register_date = request.POST['register_date'] if request.POST['register_date'] != "" else None
    marker = request.POST['marker'].strip()
    serial_no = request.POST['serial_no'].strip()
    nose = request.POST['nose'].strip()
    max_speed = request.POST['max_speed'].strip()
    drive_type = request.POST['drive_type'].strip()
    lubrication = request.POST['lubrication'].strip()
    condition = request.POST['condition'].strip()
    sp_new = SplindlePart(machine=machine,model=model,amount=amount,register_date=register_date,marker=marker,serial_no=serial_no,nose=nose,max_speed=max_speed,drive_type=drive_type,lubrication=lubrication,condition=condition)
    sp_new.save()
    return redirect('/new/sp/')

def new_task_save(request):
    type = request.POST['type'].strip()
    name = request.POST['name'].strip()
    note = request.POST['note']
    task_new = Task(type=type,name=name,note=note)
    task_new.save()
    return redirect('/new/task/')

def new_ven_save(request):
    code = request.POST['code'].strip()
    name = request.POST['name'].strip()
    address = request.POST['address']
    email = request.POST['email'].strip()
    phone_no = request.POST['phone_no'].strip()
    note = request.POST['note']
    ven_new = Vendor(code=code,name=name,address=address,email=email,phone_no=phone_no,note=note)
    ven_new.save()
    return redirect('/new/ven/')

def new_cat_save(request):
    name = request.POST['name'].strip()
    cat_new = Category(name=name)
    cat_new.save()
    return redirect('/new/cat/')

def new_sub_cat_save(request):
    name = request.POST['name'].strip()
    cat_id = request.POST['cat_id']
    description = request.POST['description']
    cat = Category.objects.get(id=cat_id)
    sub_cat_new = SubCategory(name=name,cat=cat,description=description)
    sub_cat_new.save()
    return redirect('/new/sub_cat/')

def new_mg_save(request):
    sg_name = request.POST['sg_name']
    username = request.POST['username']
    is_cc = True if request.POST.get('is_cc', False) == 'on' else False
    sg = SectionGroup.objects.get(name=sg_name)
    user = User.objects.get(username=username)
    mg_new = MailGroup(sg=sg,user=user,is_cc=is_cc)
    mg_new.save()
    return redirect('/new/mg/')

def new_pwst_save(request):
    name = request.POST['name'].strip()
    password = request.POST['password'].strip()
    note = request.POST['note']
    if request.user.employee.enable_pwst:
        pwst_new = PasswordStorage(name=name,note=note)
        pwst_new.save()
        pwi_new = PasswordItem(pwst=pwst_new,password=password)
        pwi_new.save()
    return redirect('/new/pwst/')

def edit_mc_save(request):
    is_active = True if request.POST.get('is_active', False) == 'on' else False
    mc_no = request.POST['mc_no']
    mcg_id = request.POST['mcg_id']
    sap_mc_no = request.POST['sap_mc_no'].strip()
    type = request.POST['type'].strip()
    section = request.POST['section'].strip()
    register_no = request.POST['register_no'].strip()
    asset_no = request.POST['asset_no'].strip()
    serial_no = request.POST['serial_no'].strip()
    manufacture = request.POST['manufacture'].strip()
    model = request.POST['model'].strip()
    plant = request.POST['plant'].strip()
    power = request.POST['power'].strip()
    location = request.POST['location'].strip()
    install_date = request.POST['install_date'] if request.POST['install_date'] != "" else None
    capacity = request.POST['capacity']
    note = request.POST['note']
    mcg = None
    if mcg_id != '-1':
        mcg = MachineGroup.objects.get(id=mcg_id)
    mc = Machine.objects.get(mc_no=mc_no)
    mc.is_active = is_active
    mc.mcg = mcg
    mc.sap_mc_no = sap_mc_no
    mc.type = type
    mc.section = section
    mc.register_no = register_no
    mc.asset_no = asset_no
    mc.serial_no = serial_no
    mc.manufacture = manufacture
    mc.model = model
    mc.plant = plant
    mc.power = power
    mc.location = location
    mc.install_date = install_date
    mc.capacity = capacity
    mc.note = note
    mc.save()
    return redirect('/edit/mc/' + mc.mc_no)

def edit_cp_save(request):
    id = request.POST['id']
    mat_code = request.POST['mat_code'].strip()
    amount = int(request.POST['amount'].strip())
    minimum = int(request.POST['minimum'].strip())
    note = request.POST['note']
    cp = CriticalPart.objects.get(id=id)
    cp.mat_code = mat_code
    cp.amount = amount
    cp.minimum = minimum
    cp.note = note
    cp.save()
    return redirect('/edit/cp/' + str(cp.id))

def edit_sp_save(request):
    id = request.POST['id']
    machine = request.POST['machine'].strip()
    model = request.POST['model'].strip()
    amount = int(request.POST['amount'].strip())
    register_date = request.POST['register_date'] if request.POST['register_date'] != "" else None
    marker = request.POST['marker'].strip()
    serial_no = request.POST['serial_no'].strip()
    nose = request.POST['nose'].strip()
    max_speed = request.POST['max_speed'].strip()
    drive_type = request.POST['drive_type'].strip()
    lubrication = request.POST['lubrication'].strip()
    condition = request.POST['condition'].strip()
    sp = SplindlePart.objects.get(id=id)
    sp.machine = machine
    sp.model = model
    sp.amount = amount
    sp.register_date = register_date
    sp.marker = marker
    sp.serial_no = serial_no
    sp.nose = nose
    sp.max_speed = max_speed
    sp.drive_type = drive_type
    sp.lubrication = lubrication
    sp.condition = condition
    sp.save()
    return redirect('/edit/sp/' + str(sp.id))

def edit_task_save(request):
    is_active = True if request.POST.get('is_active', False) == 'on' else False
    id = request.POST['id']
    note = request.POST['note']
    task = Task.objects.get(id=id)
    task.is_active = is_active
    task.note = note
    task.save()
    return redirect('/edit/task/' + str(task.id))

def edit_ven_save(request):
    is_active = True if request.POST.get('is_active', False) == 'on' else False
    code = request.POST['code']
    address = request.POST['address']
    email = request.POST['email']
    phone_no = request.POST['phone_no']
    note = request.POST['note']
    ven = Vendor.objects.get(code=code)
    ven.is_active = is_active
    ven.address = address
    ven.email = email
    ven.phone_no = phone_no
    ven.note = note
    ven.save()
    return redirect('/edit/ven/' + str(ven.code))

def edit_pwst_save(request):
    id = request.POST['id']
    note = request.POST['note'].strip()
    password = request.POST['password'].strip()
    pwst = PasswordStorage.objects.get(id=id)
    if request.user.employee.enable_pwst:
        pwst.note = note
        pwst.save()
        if password != None and password != "":
            pwi_new = PasswordItem(pwst=pwst,password=password)
            pwi_new.save()
    return redirect('/edit/pwst/' + str(pwst.id))

def file_save(request):
    file = request.FILES['file']
    token = request.POST['token']
    fs = FileSystemStorage()
    path = 'temp/' + token + '/' + file.name
    fs.save(path, file)
    data = {
    }
    return JsonResponse(data)

##################################### GET ######################################

def validate_old_password(request):
    username = request.GET['username']
    old_password = request.GET['old_password']
    isCorrect = True
    user = authenticate(request,username=username,password=old_password)
    if user == None:
        isCorrect = False
    data = {
        'isCorrect': isCorrect,
    }
    return JsonResponse(data)

def validate_username(request):
    username = request.GET['username']
    canUse = True
    isExist = User.objects.filter(username=username).exists()
    if isExist:
        canUse = False
    data = {
        'canUse': canUse,
    }
    return JsonResponse(data)

def validate_mc_no(request):
    mc_no = request.GET['mc_no'].strip()
    canUse = True
    isExist = Machine.objects.filter(mc_no__iexact=mc_no).exists()
    if isExist:
        canUse = False
    data = {
        'canUse': canUse,
    }
    return JsonResponse(data)

def validate_cp(request):
    name = request.GET['name'].strip()
    canUse = True
    isExist = CriticalPart.objects.filter(name__iexact=name).exists()
    if isExist:
        canUse = False
    data = {
        'canUse': canUse,
    }
    return JsonResponse(data)

def validate_task(request):
    type = request.GET['type'].strip()
    name = request.GET['name'].strip()
    canUse = True
    isExist = Task.objects.filter(name__iexact=name,type__iexact=type).exists()
    if isExist:
        canUse = False
    data = {
        'canUse': canUse,
    }
    return JsonResponse(data)

def validate_vendor_code(request):
    code = request.GET['code'].strip()
    canUse = True
    isExist = Vendor.objects.filter(code__iexact=code).exists()
    if isExist:
        canUse = False
    data = {
        'canUse': canUse,
    }
    return JsonResponse(data)

def validate_category_name(request):
    name = request.GET['name'].strip()
    canUse = True
    isExist = Category.objects.filter(name__iexact=name).exists()
    if isExist:
        canUse = False
    data = {
        'canUse': canUse,
    }
    return JsonResponse(data)

def validate_sub_category_name(request):
    name = request.GET['name'].strip()
    cat_id = request.GET['cat_id']
    canUse = True
    cat = Category.objects.get(id=cat_id)
    isExist = SubCategory.objects.filter(name__iexact=name,cat=cat).exists()
    if isExist:
        canUse = False
    data = {
        'canUse': canUse,
    }
    return JsonResponse(data)

def validate_user_in_mail_group(request):
    sg_name = request.GET['sg_name']
    username = request.GET['username']
    canUse = True
    sg = SectionGroup.objects.get(name=sg_name)
    user = User.objects.get(username=username)
    isExist = MailGroup.objects.filter(sg=sg,user=user).exists()
    if isExist:
        canUse = False
    data = {
        'canUse': canUse,
    }
    return JsonResponse(data)

def validate_pwst(request):
    name = request.GET['name'].strip()
    canUse = True
    isExist = PasswordStorage.objects.filter(name__iexact=name).exists()
    if isExist:
        canUse = False
    data = {
        'canUse': canUse,
    }
    return JsonResponse(data)

def find_emp_info(request):
    emp_id = request.GET['emp_id']
    found = False
    name = ""
    section = ""
    email = ""
    phone_no = ""
    reqs = Request.objects.filter(emp_id=emp_id).order_by('-date_published')
    if reqs:
        found = True
        name = reqs[0].name
        section = reqs[0].section
        email = reqs[0].email
        phone_no = reqs[0].phone_no
    data = {
        'found': found,
        'name': name,
        'section': section,
        'email': email,
        'phone_no': phone_no,
    }
    return JsonResponse(data)

def accept_request(request):
    req_id = request.GET['req_id']
    mc_no = request.GET['mc_no'] if request.GET['mc_no'] != 'Select' else None
    username_list = request.GET.getlist('username_list[]')
    ven_code_list = request.GET.getlist('ven_code_list[]')
    sub_cat_id_list = request.GET.getlist('sub_cat_id_list[]')
    req = Request.objects.get(id=req_id)
    if mc_no != None:
        mc = Machine.objects.get(mc_no=mc_no)
        req.mc = mc
    for username in username_list:
        user = User.objects.get(username=username)
        member_new = Member(req=req,user=user)
        member_new.save()
    for ven_code in ven_code_list:
        ven = Vendor.objects.get(code=ven_code)
        req_ven_new = RequestVendor(req=req,ven=ven)
        req_ven_new.save()
    for sub_cat_id in sub_cat_id_list:
        sub_cat = SubCategory.objects.get(id=sub_cat_id)
        req_sub_cat_new = RequestSubCategory(req=req,sub_cat=sub_cat)
        req_sub_cat_new.save()
    req.status = 'In Progress'
    req.save()
    data = {
    }
    return JsonResponse(data)

def reject_request(request):
    req_id = request.GET['req_id']
    is_nms = True if request.GET['is_nms'] == 'true' else False
    reject_reason = request.GET['reject_reason']
    sg_name = request.GET['sg_name']
    req = Request.objects.get(id=req_id)
    if is_nms:
        sg = SectionGroup.objects.get(name=sg_name)
        req.sg = sg
        #-- Email
        subject = '[CMS] New Request #' + req.req_no
        send_to = get_mail_group(sg, False)
        cc_to = get_mail_group(sg, True)
        email_template = get_template(TEMPLATE_REQUEST)
        email_content = email_template.render({
            'type' : 'New Request',
            'req' : req,
            'host_url' : HOST_URL,
        })
        send_email(subject, email_content, send_to, cc_to)
        #-- Line
        send_line(sg.line_token, req)
    else:
        req.status = 'Rejected'
        req.reason = reject_reason
        req.finish_datetime = datetime.now()
    req.save()
    data = {
    }
    return JsonResponse(data)

def join_request(request):
    req_id = request.GET['req_id']
    req = Request.objects.get(id=req_id)
    member_new = Member(req=req,user=request.user)
    member_new.save()
    data = {
    }
    return JsonResponse(data)

def leave_request(request):
    req_id = request.GET['req_id']
    req = Request.objects.get(id=req_id)
    is_member = Member.objects.filter(req=req,user=request.user).exists()
    if is_member:
        member = Member.objects.get(req=req,user=request.user)
        member.delete()
    data = {
    }
    return JsonResponse(data)

def hold_request(request):
    req_id = request.GET['req_id']
    hold_reason = request.GET['hold_reason']
    req = Request.objects.get(id=req_id)
    req.status = 'On Hold'
    req.reason = hold_reason
    req.save()
    data = {
    }
    return JsonResponse(data)

def start_work_request(request):
    req_id = request.GET['req_id']
    req = Request.objects.get(id=req_id)
    req.status = 'In Progress'
    req.reason = None
    req.save()
    data = {
    }
    return JsonResponse(data)

def complete_request(request):
    req_id = request.GET['req_id']
    finish_datetime = datetime.now() if request.GET['finish_datetime'] == "" else request.GET['finish_datetime']
    correction = request.GET['correction']
    spare_parts = request.GET['spare_parts']
    cause = request.GET['cause']
    # corrective_action = request.GET['corrective_action']
    documents = request.GET['documents']
    req = Request.objects.get(id=req_id)
    req.status = 'Complete'
    req.reason = None
    req.correction = correction
    req.spare_parts = spare_parts
    req.cause = cause
    # req.corrective_action = corrective_action
    req.documents = documents
    req.is_breakdown = False
    req.breakdown_reason = None
    req.finish_datetime = finish_datetime
    req.save()
    data = {
    }
    return JsonResponse(data)

def cancel_request(request):
    req_id = request.GET['req_id']
    cancel_reason = request.GET['cancel_reason']
    req = Request.objects.get(id=req_id)
    req.status = 'Canceled'
    req.reason = cancel_reason
    req.is_breakdown = False
    req.breakdown_reason = None
    req.finish_datetime = datetime.now()
    req.save()
    OperatorWorkingTime.objects.filter(req=req).delete()
    VendorWorkingTime.objects.filter(req=req).delete()
    MachineDowntime.objects.filter(req=req).delete()
    Costing.objects.filter(req=req).delete()
    data = {
    }
    return JsonResponse(data)

def rework_request(request):
    req_id = request.GET['req_id']
    req = Request.objects.get(id=req_id)
    req.status = 'In Progress'
    req.reason = None
    req.save()
    data = {
    }
    return JsonResponse(data)

def set_breakdown(request):
    req_id = request.GET['req_id']
    breakdown_reason = request.GET['breakdown_reason']
    req = Request.objects.get(id=req_id)
    req.is_breakdown = True
    req.breakdown_reason = breakdown_reason
    req.save()
    data = {
    }
    return JsonResponse(data)

def remove_breakdown(request):
    req_id = request.GET['req_id']
    req = Request.objects.get(id=req_id)
    req.is_breakdown = False
    req.breakdown_reason = None
    req.save()
    data = {
    }
    return JsonResponse(data)

def manage_member(request):
    req_id = request.GET['req_id']
    username_list = request.GET.getlist('username_list[]')
    req = Request.objects.get(id=req_id)
    members = Member.objects.filter(req=req)
    members.delete()
    for username in username_list:
        user = User.objects.get(username=username)
        member_new = Member(req=req,user=user)
        member_new.save()
    data = {
    }
    return JsonResponse(data)

def manage_vendor(request):
    req_id = request.GET['req_id']
    ven_code_list = request.GET.getlist('ven_code_list[]')
    req = Request.objects.get(id=req_id)
    req_vens = RequestVendor.objects.filter(req=req)
    req_vens.delete()
    for ven_code in ven_code_list:
        ven = Vendor.objects.get(code=ven_code)
        req_ven_new = RequestVendor(req=req,ven=ven)
        req_ven_new.save()
    data = {
    }
    return JsonResponse(data)

def manage_category(request):
    req_id = request.GET['req_id']
    sub_cat_id_list = request.GET.getlist('sub_cat_id_list[]')
    req = Request.objects.get(id=req_id)
    req_sub_cats = RequestSubCategory.objects.filter(req=req)
    req_sub_cats.delete()
    for sub_cat_id in sub_cat_id_list:
        sub_cat = SubCategory.objects.get(id=sub_cat_id)
        req_sub_cat_new = RequestSubCategory(req=req,sub_cat=sub_cat)
        req_sub_cat_new.save()
    data = {
    }
    return JsonResponse(data)

def post_comment(request):
    req_id = request.GET['req_id']
    comment_text = request.GET['comment_text']
    req = Request.objects.get(id=req_id)
    comment_new = Comment(req=req,text=comment_text,user=request.user)
    comment_new.save()
    data = {
    }
    return JsonResponse(data)

def delete_file(request):
    file_id = request.GET['file_id']
    file = File.objects.get(id=file_id)
    folder_path = 'media/request/' + file.req.req_no
    file_path = folder_path + '/' + file.file_name
    os.remove(file_path)
    if len(os.listdir(folder_path)) == 0:
        shutil.rmtree(folder_path, ignore_errors=False, onerror=None)
    file.delete()
    data = {
    }
    return JsonResponse(data)

def delete_comment(request):
    comment_id = request.GET['comment_id']
    comment = Comment.objects.get(id=comment_id)
    comment.delete()
    data = {
    }
    return JsonResponse(data)

def owt_save(request):
    req_id = request.GET['req_id']
    username_list = request.GET.getlist('username_list[]')
    start_datetime = request.GET['start_datetime']
    stop_datetime = request.GET['stop_datetime']
    req = Request.objects.get(id=req_id)
    for username in username_list:
        user = User.objects.get(username=username)
        owt_new = OperatorWorkingTime(req=req,user=user,start_datetime=start_datetime,stop_datetime=stop_datetime)
        owt_new.save()
    data = {
    }
    return JsonResponse(data)

def delete_owt(request):
    owt_id = request.GET['owt_id']
    owt = OperatorWorkingTime.objects.get(id=owt_id)
    owt.delete()
    data = {
    }
    return JsonResponse(data)

def vwt_save(request):
    req_id = request.GET['req_id']
    ven_code_list = request.GET.getlist('ven_code_list[]')
    start_datetime = request.GET['start_datetime']
    stop_datetime = request.GET['stop_datetime']
    req = Request.objects.get(id=req_id)
    for ven_code in ven_code_list:
        ven = Vendor.objects.get(code=ven_code)
        owt_new = VendorWorkingTime(req=req,ven=ven,start_datetime=start_datetime,stop_datetime=stop_datetime)
        owt_new.save()
    data = {
    }
    return JsonResponse(data)

def delete_vwt(request):
    vwt_id = request.GET['vwt_id']
    vwt = VendorWorkingTime.objects.get(id=vwt_id)
    vwt.delete()
    data = {
    }
    return JsonResponse(data)

def mcdt_save(request):
    req_id = request.GET['req_id']
    mc_no = request.GET['mc_no']
    start_datetime = request.GET['start_datetime']
    stop_datetime = request.GET['stop_datetime']
    req = Request.objects.get(id=req_id)
    mc = Machine.objects.get(mc_no=mc_no)
    mcdt_new = MachineDowntime(req=req,mc=mc,start_datetime=start_datetime,stop_datetime=stop_datetime)
    mcdt_new.save()
    data = {
    }
    return JsonResponse(data)

def delete_mcdt(request):
    mcdt_id = request.GET['mcdt_id']
    mcdt = MachineDowntime.objects.get(id=mcdt_id)
    mcdt.delete()
    data = {
    }
    return JsonResponse(data)

def cost_save(request):
    req_id = request.GET['req_id']
    name = request.GET['name'].strip()
    price = request.GET['price']
    req = Request.objects.get(id=req_id)
    cost_new = Costing(req=req,name=name,price=price)
    cost_new.save()
    data = {
    }
    return JsonResponse(data)

def delete_cost(request):
    cost_id = request.GET['cost_id']
    cost = Costing.objects.get(id=cost_id)
    cost.delete()
    data = {
    }
    return JsonResponse(data)

def set_tot(request):
    mcg_id = request.GET['mcg_id']
    year = request.GET['year']
    month = request.GET['month']
    time = request.GET['time']
    mcg = MachineGroup.objects.get(id=mcg_id)
    tot_is_exist = TotalOperationTime.objects.filter(mcg=mcg,year=year,month=month).exists()
    if tot_is_exist:
        tot = TotalOperationTime.objects.get(mcg=mcg,year=year,month=month)
        tot.delete()
    tot_new = TotalOperationTime(mcg=mcg,year=year,month=month,time=time)
    tot_new.save()
    data = {
    }
    return JsonResponse(data)

def set_ewt(request):
    mcg_id = request.GET['mcg_id']
    year = request.GET['year']
    month = request.GET['month']
    time = request.GET['time']
    mcg = MachineGroup.objects.get(id=mcg_id)
    ewts = EstimateWorkingTime.objects.filter(mcg=mcg,year__gt=year) | EstimateWorkingTime.objects.filter(mcg=mcg,year=year,month__gte=month)
    ewts.delete()
    ewt_new = EstimateWorkingTime(mcg=mcg,year=year,month=month,time=time)
    ewt_new.save()
    data = {
    }
    return JsonResponse(data)

def set_target(request):
    mcg_id = request.GET['mcg_id']
    year = request.GET['year']
    month = request.GET['month']
    type = request.GET['type']
    time = request.GET['time']
    mcg = MachineGroup.objects.get(id=mcg_id)
    qots = QualityObjectiveTarget.objects.filter(mcg=mcg,year__gt=year,type=type) | QualityObjectiveTarget.objects.filter(mcg=mcg,year=year,month__gte=month,type=type)
    qots.delete()
    qot_new = QualityObjectiveTarget(mcg=mcg,year=year,month=month,type=type,time=time)
    qot_new.save()
    data = {
    }
    return JsonResponse(data)

################################# File Reader ##################################

def update_machine():
    wb = load_workbook(filename = 'media/MC.xlsx')
    ws = wb.active
    skip_count = 2
    for i in range(ws.max_row + 1):
        if i < skip_count:
            continue
        mc_no = (ws['C' + str(i)].value).strip()
        section = (ws['A' + str(i)].value).strip()
        mcg_no = str((ws['B' + str(i)].value)).strip() if ws['B' + str(i)].value != None else ""
        mcg = MachineGroup.objects.get(name=mcg_no) if mcg_no != "" else None
        is_active = ws['D' + str(i)].value
        register_no = str((ws['E' + str(i)].value)).strip() if ws['E' + str(i)].value != None else ""
        asset_no = str((ws['F' + str(i)].value)).strip() if ws['F' + str(i)].value != None else ""
        manufacture = str((ws['G' + str(i)].value)).strip() if ws['G' + str(i)].value != None else ""
        plant = str((ws['H' + str(i)].value)).strip() if ws['H' + str(i)].value != None else ""
        model = str((ws['I' + str(i)].value)).strip() if ws['I' + str(i)].value != None else ""
        serial_no = str((ws['J' + str(i)].value)).strip() if ws['J' + str(i)].value != None else ""
        capacity = str((ws['K' + str(i)].value)).strip() if ws['K' + str(i)].value != None else ""
        power = str((ws['L' + str(i)].value)).strip() if ws['L' + str(i)].value != None else ""
        install_date = None if ws['M' + str(i)].value == "" else ws['M' + str(i)].value
        note = str((ws['N' + str(i)].value)).strip() if ws['N' + str(i)].value != None else ""
        sap_mc_no = str((ws['O' + str(i)].value)).strip() if ws['O' + str(i)].value != None else ""
        location = str((ws['P' + str(i)].value)).strip() if ws['P' + str(i)].value != None else ""
        type = str((ws['Q' + str(i)].value)).strip() if ws['Q' + str(i)].value != None else ""

        if Machine.objects.filter(mc_no=mc_no).exists():
            mc = Machine.objects.get(mc_no=mc_no)
            mc.section = section
            mc.mcg = mcg
            mc.is_active = is_active
            mc.register_no = register_no
            mc.asset_no = asset_no
            mc.manufacture = manufacture
            mc.plant = plant
            mc.model = model
            mc.serial_no = serial_no
            mc.capacity = capacity
            mc.power = power
            mc.install_date = install_date
            mc.note = note
            mc.sap_mc_no = sap_mc_no
            mc.location = location
            mc.type = type
            mc.save()
        else:
            new_mc = Machine(mc_no=mc_no,section=section,mcg=mcg,is_active=is_active,register_no=register_no,asset_no=asset_no,manufacture=manufacture,plant=plant,model=model,serial_no=serial_no,capacity=capacity,power=power,install_date=install_date,note=note,sap_mc_no=sap_mc_no,location=location,type=type)
            new_mc.save()
    return

def custom_script():
    reqs = Request.objects.filter(status='Canceled')
    for req in reqs:
        OperatorWorkingTime.objects.filter(req=req).delete()
        VendorWorkingTime.objects.filter(req=req).delete()
        MachineDowntime.objects.filter(req=req).delete()
        Costing.objects.filter(req=req).delete()
    return

################################ Other Function ################################

def add_front_zero(str):
    if len(str) == 1:
        return "0" + str
    return str

def create_req_no(id):
    req_no = str(id)
    for i in range(7 - len(str(id))):
        req_no = "0" + req_no
    req_no = "CMS" + req_no
    return req_no

def sort_user_by_section(users):
    f_users = []
    emps = Employee.objects.filter(user__in=users).order_by('section')
    for emp in emps:
        f_users.append(emp.user)
    return f_users

def get_set_mc(mcs):
    set_mc = []
    temp = ""
    for mc in mcs:
        if temp != mc.section:
            temp = mc.section
            set_mc.append(mc.section)
        else:
            set_mc.append(None)
    return set_mc

def get_set_task(tasks):
    set_task = []
    temp = ""
    for task in tasks:
        if temp != task.type:
            temp = task.type
            set_task.append(task.type)
        else:
            set_task.append(None)
    return set_task

def get_set_user(users):
    set_user = []
    temp = ""
    for user in users:
        if temp != user.employee.section:
            temp = user.employee.section
            set_user.append(user.employee.section)
        else:
            set_user.append(None)
    return set_user

def get_set_sc(sub_cats):
    set_sc = []
    temp = ""
    for sub_cat in sub_cats:
        if temp != sub_cat.cat.name:
            temp = sub_cat.cat.name
            set_sc.append(sub_cat.cat.name)
        else:
            set_sc.append(None)
    return set_sc

def get_set_rsc(req_sub_cats):
    set_rsc = []
    temp = ""
    for req_sub_cat in req_sub_cats:
        if temp != req_sub_cat.sub_cat.cat.name:
            temp = req_sub_cat.sub_cat.cat.name
            set_rsc.append(req_sub_cat.sub_cat.cat.name)
        else:
            set_rsc.append(None)
    return set_rsc

def get_mail_group(sg, is_cc):
    list = []
    mgs = MailGroup.objects.filter(sg=sg,is_cc=is_cc)
    for mg in mgs:
        if mg.user.is_active:
            list.append(mg.user.email)
    return list

def get_is_members(reqs, request):
    is_members = []
    for req in reqs:
        if req.status == 'Rejected':
            is_members.append(None)
        else:
            is_member = Member.objects.filter(req=req,user=request.user).exists()
            if is_member:
                is_members.append(True)
            else:
                is_members.append(False)
    return is_members

def get_has_wts(reqs):
    has_wts = []
    for req in reqs:
        if req.status == 'Rejected':
            has_wts.append(None)
        else:
            has_owt = OperatorWorkingTime.objects.filter(req=req).exists()
            has_vwt = VendorWorkingTime.objects.filter(req=req).exists()
            if has_owt or has_vwt:
                has_wts.append(True)
            else:
                has_wts.append(False)
    return has_wts

def get_has_mcdts(reqs):
    has_mcdts = []
    for req in reqs:
        if req.mc == None or req.status == 'Rejected':
            has_mcdts.append(None)
        else:
            has_mcdt = MachineDowntime.objects.filter(req=req).exists()
            if has_mcdt:
                has_mcdts.append(True)
            else:
                has_mcdts.append(False)
    return has_mcdts

def get_selected_members(req, users):
    selected_members = []
    for user in users:
        is_member_exist = Member.objects.filter(req=req,user=user).exists()
        if is_member_exist:
            selected_members.append(True)
        else:
            selected_members.append(False)
    return selected_members

def get_selected_vendors(req, vens):
    selected_vendors = []
    for ven in vens:
        is_ven_exist = RequestVendor.objects.filter(req=req,ven=ven).exists()
        if is_ven_exist:
            selected_vendors.append(True)
        else:
            selected_vendors.append(False)
    return selected_vendors

def get_selected_sc(req, sub_cats):
    selected_sc = []
    for sub_cat in sub_cats:
        is_sub_cat_exist = RequestSubCategory.objects.filter(req=req,sub_cat=sub_cat).exists()
        if is_sub_cat_exist:
            selected_sc.append(True)
        else:
            selected_sc.append(False)
    return selected_sc

def is_in_section_group(request):
    sgs = SectionGroup.objects.all()
    is_in = False
    for sg in sgs:
        if sg.name == request.user.employee.section:
            is_in = True
    return is_in

def get_years():
    MIN_SIZE = 10
    START_YEAR = 2022
    years = [str(START_YEAR)]
    temp_year = START_YEAR
    while temp_year < int(datetime.today().strftime('%Y')) or len(years) < MIN_SIZE:
        temp_year = temp_year + 1
        years.append(str(temp_year))
    return years

def get_qot(mcg, type, year, month):
    time = 0
    qots = QualityObjectiveTarget.objects.filter(mcg=mcg,type=type)
    if not qots or (not qots.filter(year=year,month__lte=month) and not qots.filter(year__lt=year)):
        time = 0
    elif not qots.filter(year=year) or not qots.filter(year=year,month__lte=month):
        qots = qots.filter(year__lt=year).order_by('-year','-month')
        time = qots[0].time
    else:
        qots = qots.filter(year=year,month__lte=month).order_by('-month')
        time = qots[0].time
    return time

def get_ewt(mcg, year, month):
    time = 0
    ewts = EstimateWorkingTime.objects.filter(mcg=mcg)
    if not ewts or (not ewts.filter(year=year,month__lte=month) and not ewts.filter(year__lt=year)):
        time = 0
    elif not ewts.filter(year=year) or not ewts.filter(year=year,month__lte=month):
        ewts = ewts.filter(year__lt=year).order_by('-year','-month')
        time = ewts[0].time
    else:
        ewts = ewts.filter(year=year,month__lte=month).order_by('-month')
        time = ewts[0].time
    return time

def get_total_cost(costs):
    result = 0
    for cost in costs:
        result = result + cost.price
    return result

def get_is_not_enough_cp(cps):
    for cp in cps:
        if (cp.minimum != 0 and cp.amount < cp.minimum) or (cp.minimum == 0 and cp.amount == 0):
            return True
    return False

def get_is_not_enough_sp(sps):
    for sp in sps:
        if sp.amount == 0:
            return True
    return False

##################################### API ####################################

def api_get_emp_work_time(request, emp_id, start_date, end_date):
    work_time = 0
    is_emp_exist = User.objects.filter(username=emp_id)
    if is_emp_exist:
        work_time = 99
    result = {
        'work_time': work_time,
    }
    data = json.dumps(result)
    return HttpResponse(data, content_type='application/json')

##################################### Email ####################################

class EmailThread(threading.Thread):

    def __init__(self, email):
        self.email = email
        threading.Thread.__init__(self)

    def run(self):
        self.email.send(fail_silently=False)

def send_email(subject, email_content, send_to, cc_to):
    try:
        email = EmailMessage(subject, email_content, EMAIL_HOST_USER, send_to, cc_to)
        email.content_subtype = "html"
        EmailThread(email).start()
    except Exception:
        traceback.print_exc()
    return

def send_line(token, req):
    #-- Line
    line = Line(token)
    line.send_msg(f'\n New Request: {req.req_no}\n Request Link: {HOST_URL}req/{str(req.id)}\n Request By:{str(req.emp_id)} | {req.name} ({req.section})\n Phone No: {req.phone_no}\n Description: {req.description}')
    return